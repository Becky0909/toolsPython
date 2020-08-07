from datetime import datetime
import openpyxl
from decimal import Decimal, getcontext


# 解析文件:queryType=1：total_click表，2：service_number表
def analyseDataHW(queryType):
    # 批量插入数据库的数据列表
    data = []
    # 文件覆盖的日期段
    dateSet = set()
    book = openpyxl.load_workbook("../sql_app/test.xlsx")
    # total_click表
    if queryType == 1:
        sheets = book.sheetnames
        for j in range(len(sheets)):  # 遍历表的sheet
            table = book[sheets[j]]
            for i in range(2, table.max_row+1):
                date=table.cell(i, 1).value
                time = datetime.strptime(date, '%Y-%m-%d')
                year = time.date().year
                month = time.date().month
                day = time.date().day
                date = year*10000+month*100+day
                industry=str(table.cell(i, 2).value)
                merchantName=str(table.cell(i, 3).value)
                locatedField=str(table.cell(i, 4).value)
                totalClick=int(table.cell(i, 5).value)
                values = (industry, merchantName, locatedField, totalClick, date)
                data.append(values)
                dateSet.add(date)
    # service_number表
    else:
        sheets = book.sheetnames
        for j in range(len(sheets)):
            table = book.get_sheet_by_name(sheets[j])
            for i in range(2, table.max_row+1):
                date = table.cell(i, 1).value
                time = datetime.strptime(date, '%Y-%m-%d')
                year = time.date().year
                month = time.date().month
                day = time.date().day
                date = year * 10000 + month * 100 + day
                serviceName = str(table.cell(i, 2).value)
                industry = str(table.cell(i, 3).value)
                merchantName = str(table.cell(i, 4).value)
                locatedField = str(table.cell(i, 5).value)
                openServiceNum = str(table.cell(i, 6).value)
                openMenuNum = str(table.cell(i, 7).value)
                openPageNum = str(table.cell(i, 8).value)
                pageDayPV = int(table.cell(i, 9).value)
                pageDayUV = int(table.cell(i, 10).value)
                pageQuickAppPV = int(table.cell(i, 11).value)
                pageQuickAppUV = int(table.cell(i, 12).value)
                chatPV = int(table.cell(i, 13).value)
                chatUV = int(table.cell(i, 14).value)
                values = (serviceName, industry, merchantName,
                          locatedField, openServiceNum, openMenuNum,
                          openPageNum, pageDayPV, pageDayUV, pageQuickAppPV,
                          pageQuickAppUV, chatPV, chatUV, date)
                data.append(values)
                dateSet.add(date)
    return data, dateSet


def analyseDataXM():
    data1 = []
    data2 = []
    dateSet = set()
    book = openpyxl.load_workbook("../sql_app/test.xlsx")
    sheets = book.sheetnames
    for j in range(len(sheets)):
        table = book[sheets[j]]
        for i in range(6, table.max_row):
            date = table.cell(i, 3).value
            time = datetime.strptime(date, '%Y-%m-%d')
            year = time.date().year
            month = time.date().month
            day = time.date().day
            date = year * 10000 + month * 100 + day
            name = str(table.cell(i, 2).value).split('-')
            expo = int(table.cell(i, 4).value)
            click = int(table.cell(i, 5).value)
            download = int(table.cell(i, 6).value)
            getcontext().prec = 9
            clickRate = str(table.cell(i, 7).value)
            clickRate = Decimal(clickRate[:-1]).quantize(Decimal('0.0000'))
            values = (name[1], expo, click,
                      download, clickRate, date)
            # 按钮数据
            if name[5] == "按钮":
                data1.append(values)
            # 菜单数据
            else:
                data2.append(values)
            # 记录此表所含日期，插入时覆盖该日期范围内数据
            dateSet.add(date)
        return data1, data2, dateSet