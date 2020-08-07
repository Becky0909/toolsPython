from datetime import datetime
import openpyxl
import err_model


# 校验列数：total_click表5列，service_number表14列，小米表单7列
def rowExam(type):
    ref = {1:5, 2:14, 3:7}
    book = openpyxl.load_workbook("../sql_app/test.xlsx")
    sheets = book.sheetnames
    # 遍历表的sheet
    for j in range(len(sheets)):
        table = book[sheets[j]]
        # 华为表直接解析列数
        if type == 1 or type == 2:
            if table.max_column != ref[type]:
                return False
        # 小米表用max_column方法容易解析错误，检验第八列是否为空
        else:
            if table.cell(5, ref[type] + 1).value is not None:
                return False
    return True


# 校验数据：queryType=1：total_click表，2：service_number，3：小米表单
# 检验数据是否为空、类型是否正确、长度范围正确
def formExam(queryType):
    data = []
    # flagGlobal=1所有数据正确，0存在错误
    flagGlobal = 1
    book = openpyxl.load_workbook("../sql_app/test.xlsx")
    if queryType == 1:
        sheets = book.sheetnames
        for j in range(len(sheets)):
            table = book[sheets[j]]
            # 逐行检验
            for i in range(2, table.max_row):
                # 第i行
                rowDic = {'rowNumber': i}
                # flagLocal=1该行数据正确，0存在错误
                flagLocal = 1
                # 第一列日期校验
                date = table.cell(i, 1).value
                if date and type(date) == str:
                    try:
                        time = datetime.strptime(date, '%Y-%m-%d')
                        year = time.date().year
                        month = time.date().month
                        day = time.date().day
                        # 检验年月日不为0
                        tmp = 1/year + 1/month + 1/day
                        rowDic.update({'dataTime': err_model.dataModel(date, 1)})
                    except Exception:
                        rowDic.update({'dataTime': err_model.dataModel(date, 2)})
                        flagLocal = 0
                        flagGlobal = 0
                else:
                    rowDic.update({'dataTime': err_model.dataModel(date, 2)})
                    flagLocal = 0
                    flagGlobal = 0
                names = ['industry', 'merchantName', 'locatedField']
                # 2-4列校验
                for index in range(2, 5):
                    rowName = table.cell(i, index).value
                    if rowName and type(rowName) == str and len(rowName) <= 127:
                        rowDic.update({names[index-2]: err_model.dataModel(rowName, 1)})
                    else:
                        rowDic.update({names[index-2]: err_model.dataModel(rowName, 2)})
                        flagGlobal = 0
                        flagLocal = 0
                # 第5列校验
                totalClick = table.cell(i, 5).value
                if totalClick and str(totalClick).isdigit() and len(str(totalClick)) <= 20:
                    rowDic.update({'totalClick': err_model.dataModel(totalClick, 1)})
                else:
                    rowDic.update({'totalClick': err_model.dataModel(totalClick, 2)})
                    flagGlobal = 0
                    flagLocal = 0

                if flagLocal == 0:  # 该行有错
                    data.append(rowDic)
                    if len(data) == 30:  # 只显示30条
                        break

    elif queryType == 2:
        sheets = book.sheetnames
        for j in range(len(sheets)):
            table = book[sheets[j]]
            for i in range(2, table.max_row):
                rowDic = {'rowNumber': i}
                flagLocal = 1
                # 第一列日期校验
                date = table.cell(i, 1).value
                if date and type(date) == str:
                    try:
                        time = datetime.strptime(date, '%Y-%m-%d')
                        year = time.date().year
                        month = time.date().month
                        day = time.date().day
                        tmp = 1/year + 1/month + 1/day
                        rowDic.update({'dataTime':err_model.dataModel(date, 1)})
                    except Exception:
                        rowDic.update({'dataTime': err_model.dataModel(date, 2)})
                        flagLocal = 0
                        flagGlobal = 0
                else:
                    rowDic.update({'dataTime': err_model.dataModel(date, 2)})
                    flagLocal = 0
                    flagGlobal = 0
                names = ['serviceName', 'industry', 'merchantName', 'locatedField']
                # 2-5列校验
                for index in range(2, 6):
                    rowName = table.cell(i, index).value
                    if rowName and type(rowName) == str and len(rowName) <= 127:
                        rowDic.update({names[index - 2]: err_model.dataModel(rowName, 1)})
                    else:
                        rowDic.update({names[index - 2]: err_model.dataModel(rowName, 2)})
                        flagGlobal = 0
                        flagLocal = 0
                names = ['openServiceNum', 'openMenuNum', 'openPageNum', 'pageDayPV', 'pageDayUV', 'pageQuickAppPV',
                         'pageQuickAppUV', 'chatPV', 'chatUV']
                # 6-14列校验
                for index in range(6, 15):
                    rowName = str(table.cell(i, index).value)
                    if rowName and rowName.isdigit() and len(rowName) <= 20:
                        rowDic.update({names[index - 6]: err_model.dataModel(rowName, 1)})
                    else:
                        rowDic.update({names[index - 6]: err_model.dataModel(rowName, 2)})
                        flagGlobal = 0
                        flagLocal = 0
                if flagLocal == 0:  # 该行有错
                    data.append(rowDic)
                    if len(data) == 30:  # 只显示30条
                        break
    # 小米表格数据校验
    else:
        sheets = book.sheetnames
        for j in range(len(sheets)):
            table = book.get_sheet_by_name(sheets[j])
            for i in range(6, table.max_row):
                rowDic = {'rowNumber': i}
                flagLocal = 1
                nameRaw = str(table.cell(i, 2).value)
                name = nameRaw.split('-')
                # 日期列的判断条件
                tmp = bool(len(name) >= 6 and type(name[1]) == str and len(name[1]) <= 127 and name[5] in ["按钮", "菜单"])
                if tmp:
                    rowDic.update({'serviceName': err_model.dataModel(nameRaw, 1)})
                else:
                    rowDic.update({'serviceName': err_model.dataModel(nameRaw, 2)})
                    flagLocal = 0
                    flagGlobal = 0
                date = table.cell(i, 3).value
                if date and type(date) == str:
                    try:
                        time = datetime.strptime(date, '%Y-%m-%d')
                        year = time.date().year
                        month = time.date().month
                        day = time.date().day
                        tmp = 1/year + 1/month + 1/day
                        rowDic.update({'dataTime': err_model.dataModel(date, 1)})
                    except Exception:
                        rowDic.update({'dataTime': err_model.dataModel(date, 2)})
                        flagLocal = 0
                        flagGlobal = 0
                else:
                    rowDic.update({'dataTime': err_model.dataModel(date, 2)})
                    flagLocal = 0
                    flagGlobal = 0

                names = ['expo', 'click', 'download']
                for index in range(4, 7):
                    rowName = str(table.cell(i, index).value)
                    if rowName and rowName.isdigit() and len(rowName) <= 20:
                        rowDic.update({names[index - 4]: err_model.dataModel(rowName, 1)})
                    else:
                        rowDic.update({names[index - 4]: err_model.dataModel(rowName, 2)})
                        flagGlobal = 0
                        flagLocal = 0

                # 去除百分号获取数字的字符串形式
                clickRate = str(table.cell(i, 7).value)[:-2]
                # 带小数点，小数点不超过4位，总共不超过9位
                if clickRate.count('.') == 1:
                    left = clickRate.split('.')[0]
                    right = clickRate.split('.')[1]
                    tmp = bool(right.isdigit() and left.isdigit() and len(right) <= 4 and len(right)+len(left) <= 9)
                # 不带小数点，不超过9位
                else:
                    tmp = bool(clickRate.isdigit() and len(clickRate) <= 9)
                if tmp:
                    rowDic.update({'clickRate': err_model.dataModel(clickRate, 1)})
                else:
                    rowDic.update({'clickRate': err_model.dataModel(clickRate, 2)})
                    flagGlobal = 0
                    flagLocal = 0
                if flagLocal == 0:  # 该行有错
                    data.append(rowDic)
                    if len(data) == 30:  # 只显示30条
                        break
    return flagGlobal, data

